"""Parser for California WARN Act XLSX workbooks (Phase 4)."""

from __future__ import annotations

import io
import re
from datetime import date, datetime
from typing import Any

import openpyxl

from cam.ingestion.warn import WarnRecord


def _normalize_header(raw: str) -> str:
    """Normalize a header by lowercasing and collapsing whitespace."""
    return re.sub(r"\s+", " ", raw.lower().strip())


def _parse_xlsx_date(value: Any) -> date | None:
    """Parse a date from an XLSX cell value.

    Handles:
    - datetime objects (from data_only=True with date formatting)
    - date objects
    - strings (MM/DD/YYYY or YYYY-MM-DD format)
    """
    if value is None:
        return None

    # If already a datetime or date object
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    # Try to parse as string
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None

        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d-%b-%Y"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue

    return None


def _parse_xlsx_employees(value: Any) -> int | None:
    """Parse employee count from an XLSX cell value.

    Handles:
    - int values
    - float values that are integral (openpyxl often types numeric cells as
      float, e.g. ``15.0``) — non-integral floats are rejected
    - strings like '1,200'
    """
    if value is None:
        return None

    # bool is a subclass of int — exclude it explicitly.
    if isinstance(value, bool):
        return None

    if isinstance(value, int):
        return value if value > 0 else None

    if isinstance(value, float):
        if value.is_integer():
            n = int(value)
            return n if n > 0 else None
        return None

    # Try to parse as string
    if isinstance(value, str):
        raw = value.strip().replace(",", "")
        try:
            n = int(raw)
            return n if n > 0 else None
        except ValueError:
            return None

    return None


class CaXlsxStructureError(ValueError):
    """Raised when the CA WARN workbook is missing its expected structure.

    Signals upstream drift (renamed sheet, missing header, dropped columns) so
    the caller records a parse *error* rather than silently ingesting 0 records
    with 0 errors — which would let CA ingestion die quietly during a workbook
    change and be masked by other sources in ``ingest --source all``.
    """


# Columns whose absence means the workbook layout changed enough that we cannot
# trust the parse (as opposed to a legitimately empty week of notices).
_REQUIRED_COLUMNS = ("company", "notice date")


def _cell(row: tuple[Any, ...], idx: int) -> Any:
    """Safely read cell *idx* from a (possibly short) row; None if out of range."""
    return row[idx] if idx < len(row) else None


def parse_ca_xlsx(content: bytes) -> list[WarnRecord]:
    """Parse a California WARN Act XLSX workbook.

    The workbook contains a sheet named 'Detailed WARN Report ' (with trailing space).
    Row 1 is a title banner; row 2 is the header; data starts at row 3.
    Header cells contain embedded newlines; they are matched by normalizing
    (lowercase, collapse whitespace).

    Rows are streamed (``read_only=True`` is honoured — the sheet is never
    materialised in full).  An empty-but-valid sheet returns ``[]``; a
    *structural* failure raises :class:`CaXlsxStructureError` so drift is loud.

    Parameters
    ----------
    content : bytes
        The raw XLSX file content.

    Returns
    -------
    list[WarnRecord]
        List of parsed WARN records.

    Raises
    ------
    CaXlsxStructureError
        If the expected sheet, header row, or required columns are missing.
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    # Find the correct sheet by normalized name
    sheet = None
    for name in wb.sheetnames:
        if _normalize_header(name) == "detailed warn report":
            sheet = wb[name]
            break

    if sheet is None:
        raise CaXlsxStructureError(
            "CA WARN workbook: 'Detailed WARN Report' sheet not found; "
            f"available sheets: {wb.sheetnames}"
        )

    # Stream rows: row 1 is the banner, row 2 the header, data from row 3.
    row_iter = sheet.iter_rows(values_only=True)
    try:
        next(row_iter)  # banner row
        header_row = next(row_iter)
    except StopIteration:
        raise CaXlsxStructureError(
            "CA WARN workbook: 'Detailed WARN Report' sheet has no header row"
        ) from None

    if not header_row or all(c is None for c in header_row):
        raise CaXlsxStructureError("CA WARN workbook: header row is empty")

    # Build column index map: normalized_header -> column_index
    col_index: dict[str, int] = {}
    for col_idx, cell_value in enumerate(header_row):
        if cell_value is not None:
            col_index[_normalize_header(str(cell_value))] = col_idx

    missing = [c for c in _REQUIRED_COLUMNS if c not in col_index]
    if missing:
        raise CaXlsxStructureError(
            f"CA WARN workbook: header missing required columns {missing}; got {sorted(col_index)}"
        )

    records: list[WarnRecord] = []
    for row_data in row_iter:
        # Stop at first fully empty row (all cells are None)
        if all(v is None for v in row_data):
            break

        # Skip rows with empty company
        raw_company = _cell(row_data, col_index["company"])
        if not raw_company:
            continue

        company = str(raw_company).strip()
        notice_date = _parse_xlsx_date(_cell(row_data, col_index["notice date"]))

        employees_affected = None
        if "no. of employees" in col_index:
            employees_affected = _parse_xlsx_employees(
                _cell(row_data, col_index["no. of employees"])
            )

        city = ""
        if "address" in col_index:
            city_val = _cell(row_data, col_index["address"])
            city = str(city_val).strip() if city_val else ""

        county = ""
        if "county/parish" in col_index:
            county_val = _cell(row_data, col_index["county/parish"])
            county = str(county_val).strip() if county_val else ""

        layoff_type = ""
        if "layoff/ closure" in col_index:
            layoff_val = _cell(row_data, col_index["layoff/ closure"])
            layoff_type = str(layoff_val).strip() if layoff_val else ""

        # Build raw dict from all columns.  Values must be JSON-serializable
        # because they are persisted into Event.raw_json (and the DLQ) — coerce
        # datetime/date cells to ISO strings rather than storing raw objects.
        raw: dict[str, Any] = {}
        for header, col_idx in col_index.items():
            cell = _cell(row_data, col_idx)
            raw[header] = cell.isoformat() if isinstance(cell, (datetime, date)) else cell

        records.append(
            WarnRecord(
                state_code="CA",
                company=company,
                notice_date=notice_date,
                employees_affected=employees_affected,
                city=city,
                county=county,
                layoff_type=layoff_type,
                raw=raw,
            )
        )

    return records
